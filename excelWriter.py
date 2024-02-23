import os
import pandas as pd
class ExcelWriter:
    def __init__(self):
        from openpyxl import Workbook
        self.wb = Workbook()

    def write_data(self, data: list):
        sheet = self.wb.active
        for row in data:
            sheet.append(row)

    def update_value(self, row:int, col:int, value):
        sheet = self.wb.active
        sheet.cell(row=row, column=col).value = value

    def add_formula(self, row:int, col:int, formula:str):
        sheet = self.wb.active
        sheet.cell(row=row, column=col).value = formula

    def save(self, path:str, name:str):
        self.wb.save(os.path.join(path, name))

    def csv_to_excel(self, filepath: str, destination: str, name: str):
        try:
            csv_converter = pd.read_csv(filepath)
            csv_converter.to_excel(os.path.join(destination, name), index=False, header=True)
        except:
            print('error converting')

if __name__ == '__main__':
    from path_root import MyPath
    excel_dir = MyPath('data')

    table = ExcelWriter()
    data = [['имя', 'количество', 'цена'],
            ['игрушки', 10, 5.5],
            ['фломастеры', 12, 6.4],
            ['диски', 15, 9.99]]

    table.write_data(data)
    table.update_value(2,2, 27)
    max_row = table.wb.active.max_row + 1
    table.update_value(max_row, 2, 'Сумма')
    #table.add_formula(max_row, 3, f'SUM(A3:{max_row-1}3')
    table.save(str(excel_dir), 'mytable.xlsx')

    csv_file = MyPath('data/bestsellers_with_categories_2022_03_27 (1).csv')
    table.csv_to_excel(str(csv_file), 'data', 'from_csv.xlsx')