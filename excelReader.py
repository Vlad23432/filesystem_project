import openpyxl
from path_root import MyPath

class ExcelReader:
    def __init__(self, path):
        self.filename = path
        self.wb = openpyxl.load_workbook(self.filename, data_only=True)

    def get_sheetnames(self):
        """
        возвращает список имен листов таблицы файла Excel
        :return: list
        """
        return self.wb.sheetnames

    def get_active_sheet(self):
        return self.wb.active
    def get_cell_value(self, sheet_name, cell_address):
        """
        Возвращает значение ячейки
        :param sheet_name: имя листа таблицы Excel
        :param cell_address:
        :return:
        """
        sheet = self.wb[sheet_name]
        return sheet[cell_address].value

    def itetate_excel_sheet(self, sheet_name):
        sheet = self.wb[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                print(cell.coordinate, cell.value)
            print()


if __name__ == "__main__":
    filepath = MyPath('./data/example.xlsx')
    print(filepath)
    reader = ExcelReader(str(filepath))

    sheetnames = reader.get_sheetnames()
    print(sheetnames) #  получаем активный лист
    active_sheet = reader.get_active_sheet()
    print(active_sheet) # print(active_sheet.title) - вернет только название

    cell_value = reader.get_cell_value(active_sheet.title, 'A1')
    print(cell_value)

    reader.itetate_excel_sheet(active_sheet.title)